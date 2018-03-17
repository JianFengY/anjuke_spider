"""
Created on 2018/3/16
@Author: Jeff Yang
"""
import requests
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook
import datetime
import re


def get_page(url):
    """获取页面源码"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.162 Safari/537.36'
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.text
    return None


def get_houses_url(html):
    """获取一页中的房子URL"""
    doc = pq(html)
    lis = doc('.list-item').items()
    urls = []
    for li in lis:
        url = li('.house-title a').attr('href')
        # yield url
        urls.append(url)
    return urls


def get_house_info(html):
    """获取房子信息"""
    doc = pq(html)
    unit_price = doc('.third-col.detail-col dl:nth-child(2) dd').text()  # 单价
    total_price = doc('.light.info-tag').text()  # 总价
    area = doc('.second-col.detail-col dl:nth-child(2) dd').text()  # 面积
    floor = doc('.second-col.detail-col dl:nth-child(4) dd').text()  # 楼层
    title = doc('.long-title').text()  # 标题
    business_circle = doc('.iconfont.icon-area').siblings().text()  # 商圈

    trend_url = doc('.iconfont.icon-area').siblings().attr('href')  # 楼盘走势页面URL
    trend_html = get_page(trend_url)
    trend_doc = pq(trend_html)
    district = trend_doc('.bigArea .curr').text()  # 区域
    avg_price = trend_doc('.highLight em').text()  # 楼盘均价
    comp_last_month = trend_doc('.trendR > h2:nth-child(3) > i:nth-child(1)').text()  # 同比上月
    comp_last_year = trend_doc('.trendR > h2:nth-child(3) > i:nth-child(2)').text()  # 同比去年

    house_type = doc('.second-col.detail-col dl:nth-child(1) dd').text().strip()  # 户型
    direction = doc('.second-col.detail-col dl:nth-child(3) dd').text()  # 朝向
    number = doc('.house-encode').text()[6:-17]  # 房源编码
    age = doc('.first-col.detail-col dl:nth-child(3) dd').text()  # 年代
    community = doc('.first-col.detail-col dl:nth-child(1) dd').text()  # 小区
    # address = doc('.first-col.detail-col dl:nth-child(2) dd').text()[:-2]  # 地址
    # downpayment = doc('.third-col.detail-col dl:nth-child(3) dd').text()  # 首付
    monthly_payment = int(float(total_price[:-1]) * 0.7 * (0.049 / 12) * (1 + (0.049 / 12)) ** 240 / (
        (1 + (0.049 / 12)) ** 240 - 1) * 10000)  # 月供
    scrape_time = datetime.datetime.now().strftime('%Y/%m/%d')  # 抓取时间

    pattern = re.compile('comm_lat.*?\'(\d+\.\d+)\',.*?comm_lng.*?\'(\d+.\d+)\'', re.S)
    item = re.findall(pattern, html)
    lat = item[0][0]  # 纬度
    lng = item[0][1]  # 经度
    return {
        'unit_price': unit_price,
        'total_price': total_price,
        'area': area,
        'floor': floor,
        'title': title,
        'business_circle': business_circle,
        'district': district,
        'avg_price': avg_price,
        'comp_last_month': comp_last_month,
        'comp_last_year': comp_last_year,
        'house_type': house_type,
        'direction': direction,
        'number': number,
        'age': age,
        'community': community,
        # 'address': address,
        # 'downpayment': downpayment,
        'monthly_payment': monthly_payment,
        'scrape_time': scrape_time,
        'lat': lat,
        'lng': lng,
    }


if __name__ == '__main__':
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'house'
    sheet['A1'] = '单价'
    sheet['B1'] = '总价'
    sheet['C1'] = '面积'
    sheet['D1'] = '楼层'
    sheet['E1'] = '标题'
    sheet['F1'] = '商圈'
    sheet['G1'] = '区域'
    sheet['H1'] = '楼盘均价'
    sheet['I1'] = '同比上月'
    sheet['J1'] = '同比去年'
    sheet['K1'] = '户型'
    sheet['L1'] = '朝向'
    sheet['M1'] = '房源编码'
    sheet['N1'] = '建造年代'
    sheet['O1'] = '小区'
    sheet['P1'] = '月供'
    sheet['Q1'] = '抓取时间'
    sheet['R1'] = '纬度'
    sheet['S1'] = '经度'
    sheet['T1'] = 'URL地址'
    wb.save('house2.xlsx')

    urls = ['https://guangzhou.anjuke.com/sale/p' + str(i) + '/#filtersort' for i in range(1, 51)]
    count = 0
    for url in urls:
        wb = load_workbook("house2.xlsx")
        sheet = wb["house"]
        html = get_page(url)
        houses_urls = get_houses_url(html)
        for i, house_url in enumerate(houses_urls):
            house_html = get_page(house_url)
            house_info = get_house_info(house_html)
            house_info['url'] = house_url
            info = list(house_info.values())
            print(info)
            sheet.append(info)
        wb.save('house2.xlsx')
        count += len(houses_urls)
