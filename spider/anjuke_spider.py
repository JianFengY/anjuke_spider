"""
Created on 2018/3/15
@Author: Jeff Yang
"""
import requests
from pyquery import PyQuery as pq
from openpyxl import Workbook, load_workbook


def get_page(url):
    """获取页面源码"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/65.0.3325.162 Safari/537.36'
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:
        return response.text
    return None


def get_houses_url(html):
    """获取一页中的房子URL"""
    doc = pq(html)
    lis = doc('.list-item').items()
    urls = []
    for li in lis:
        url = li('.house-title a').attr('href')
        # yield url
        urls.append(url)
    return urls


def get_house_info(html):
    """获取房子信息"""
    doc = pq(html)
    unit_price = doc('.third-col.detail-col dl:nth-child(2) dd').text()  # 单价
    total_price = doc('.light.info-tag').text()  # 总价
    area = doc('.second-col.detail-col dl:nth-child(2) dd').text()  # 面积
    floor = doc('.second-col.detail-col dl:nth-child(4) dd').text()  # 楼层
    title = doc('.long-title').text()  # 标题
    # url = scrapy.Field()  # URL地址
    house_type = doc('.second-col.detail-col dl:nth-child(1) dd').text().strip()  # 户型
    number = doc('.house-encode').text()[6:-17]  # 房源编码
    age = doc('.first-col.detail-col dl:nth-child(3) dd').text()  # 年代
    community = doc('.first-col.detail-col dl:nth-child(1) dd').text()  # 小区
    address = doc('.first-col.detail-col dl:nth-child(2) dd').text()[:-2]  # 地址
    downpayment = doc('.third-col.detail-col dl:nth-child(3) dd').text()  # 首付
    # monthly_payment = doc('#reference_monthpay').text()  # 月供
    return {
        'unit_price': unit_price,
        'total_price': total_price,
        'area': area,
        'floor': floor,
        'title': title,
        'house_type': house_type,
        'number': number,
        'age': age,
        'community': community,
        'address': address,
        'downpayment': downpayment,
        # 'monthly_payment': monthly_payment,
    }


if __name__ == '__main__':
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'house'
    sheet['A1'] = '单价'
    sheet['B1'] = '总价'
    sheet['C1'] = '面积'
    sheet['D1'] = '楼层'
    sheet['E1'] = '标题'
    sheet['F1'] = '户型'
    sheet['G1'] = '房源编码'
    sheet['H1'] = '建造年代'
    sheet['I1'] = '小区'
    sheet['J1'] = '地址'
    sheet['K1'] = '首付'
    sheet['L1'] = 'URL地址'
    wb.save('house.xlsx')

    urls = ['https://guangzhou.anjuke.com/sale/p' + str(i) + '/#filtersort' for i in range(1, 51)]
    count = 0
    for url in urls:
        wb = load_workbook("house.xlsx")
        sheet = wb["house"]
        html = get_page(url)
        houses_urls = get_houses_url(html)
        for i, house_url in enumerate(houses_urls):
            house_html = get_page(house_url)
            house_info = get_house_info(house_html)
            house_info['url'] = house_url
            print(house_info)
            sheet.append(list(house_info.values()))
        count += len(houses_urls)
        wb.save('house.xlsx')
        # html = get_page('https://guangzhou.anjuke.com/sale/p1/#filtersort')
        # for house_url in get_houses_url(html):
        #     house_html = get_page(house_url)
        #     house_info = get_house_info(house_html)
        #     print(house_info)
